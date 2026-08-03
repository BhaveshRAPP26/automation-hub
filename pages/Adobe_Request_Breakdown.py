# Adobe Request Breakdown Streamlit App
import urllib.parse
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl.styles import Font

st.set_page_config(layout="wide")
st.title("Adobe Analytics Request Breakdown")

def classify(p):
    m={"lnk_o":"Custom Link","lnk_d":"Download Link","lnk_e":"Exit Link"}
    if p.get("pev2","").lower() in m:return m[p["pev2"].lower()]
    return "Page View" if p.get("pageName") else "Other"

def parse(t):
    reqs=[];cur=""
    for l in [x.strip() for x in t.splitlines() if x.strip()]:
        if "/b/ss/" in l:
            if cur:reqs.append(cur)
            cur=l
        elif cur:
            cur+="&"+l
    if cur:reqs.append(cur)
    out=[]
    for r in reqs:
        d=urllib.parse.unquote(r)
        if "?" in d:d=d.split("?",1)[1]
        p={}
        for i in d.split("&"):
            if "=" in i:
                k,v=i.split("=",1);p[k]=v
        p["Request Type"]=classify(p)
        out.append(p)
    return out

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font


def xl(tables):

    wb = Workbook()
    ws = wb.active
    ws.title = "Adobe Requests"

    row = 1

    for section, df in tables.items():

        # Section title
        cell = ws.cell(row=row, column=1)
        cell.value = section.upper()
        cell.font = Font(bold=True, size=14)

        row += 1

        # Headers
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        row += 1

        # Data
        for values in df.itertuples(index=False):

            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col).value = value

            row += 1

        row += 2

    # Auto-size columns
    for column in ws.columns:
        max_length = 0

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column[0].column_letter].width = min(max_length + 3, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

txt=st.text_area("Requests",height=300)
if st.button("Process"):
    p=parse(txt)
    g={}
    for x in p:g.setdefault(x["Request Type"],[]).append(x)
    tables={k:pd.DataFrame(v) for k,v in g.items()}
    for k,d in tables.items():
        st.subheader(k);st.dataframe(d,use_container_width=True,hide_index=True)
    if tables:
        st.download_button("Download Excel",xl(tables),"Adobe_Request_Breakdown.xlsx")
