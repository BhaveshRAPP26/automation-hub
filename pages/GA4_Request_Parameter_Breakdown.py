import urllib.parse
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font

st.set_page_config(page_title="GA4 Request Parameter Breakdown", layout="wide")
st.title("GA4 Network Request Parameter Breakdown")

def parse_requests(text):
    lines = [x.strip() for x in text.splitlines() if x.strip()]
    requests = []
    current = ""

    for line in lines:
        if (
            "google-analytics.com/g/collect" in line
            or "google-analytics.com/mp/collect" in line
            or "/g/collect?" in line
            or "/mp/collect?" in line
        ):
            if current:
                requests.append(current)
            current = line
        elif current:
            current += line if current.endswith("&") or line.startswith("&") else "&" + line

    if current:
        requests.append(current)

    results = []

    for request in requests:
        decoded = urllib.parse.unquote(request)

        if "?" in decoded:
            decoded = decoded.split("?", 1)[1]

        params = {}

        for item in decoded.split("&"):
            if "=" not in item:
                continue

            key, value = item.split("=", 1)
            key = key.replace("ep.", "").replace("up.", "")
            params[key] = value

        event_name = params.get("en", "Unknown")

        result = {
            "Network Request": request,
            "Event Name": event_name
        }

        result.update(params)
        results.append(result)

    return results


def build_tables(parsed):
    grouped = {}

    for request in parsed:
        grouped.setdefault(request.get("Event Name", "Unknown"), []).append(request)

    tables = {}

    for event_name, rows in grouped.items():
        df = pd.DataFrame(rows)

        ordered = ["Network Request", "Event Name"]

        if "click_id_hit" in df.columns:
            ordered.append("click_id_hit")
        elif "click_id" in df.columns:
            ordered.append("click_id")

        ordered += sorted(c for c in df.columns if c not in ordered)

        tables[event_name] = df[ordered]

    return tables


def create_excel(tables):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GA4 Requests"

    row = 1

    for event_name, df in tables.items():
        cell = worksheet.cell(row=row, column=1, value=event_name.upper())
        cell.font = Font(bold=True, size=14)
        row += 1

        for col, name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row, column=col, value=name)
            cell.font = Font(bold=True)

        row += 1

        for values in df.itertuples(index=False):
            for col, value in enumerate(values, 1):
                worksheet.cell(row=row, column=col, value=value)
            row += 1

        row += 3

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(max_length + 3, 80)

    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


requests = st.text_area(
    "Paste GA4 Network Requests",
    height=300
)

if st.button("Process Requests"):
    parsed = parse_requests(requests)

    if not parsed:
        st.warning("No GA4 network requests were detected.")
    else:
        st.success(f"{len(parsed)} request(s) processed.")

        tables = build_tables(parsed)

        for event_name, df in tables.items():
            st.subheader(event_name)
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True
            )

        st.download_button(
            label="Download Excel",
            data=create_excel(tables),
            file_name="GA4_Request_Breakdown.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
