import urllib.parse
from io import BytesIO
from openpyxl.styles import Font
import pandas as pd
import streamlit as st

# -----------------------------------------------------------------------------
# Streamlit Configuration
# -----------------------------------------------------------------------------

st.set_page_config(page_title="GA4 Request Breakdown", layout="wide")

st.title("GA4 Network Request Breakdown")

st.write(
    "Paste one or more GA4 network requests below (one request per line). "
    "The application will decode every parameter, group requests by event name "
    "and allow exporting the results to Excel."
)

# -----------------------------------------------------------------------------
# Functions
# -----------------------------------------------------------------------------

def parse_requests(request_text):
    """
    Parse GA4 requests.

    Supports:
      • Full request on one line
      • Request split across multiple lines
      • Header + payload
      • Multiple requests pasted together
    """

    parsed_requests = []

    lines = [
        line.strip()
        for line in request_text.splitlines()
        if line.strip()
    ]

    requests = []

    current_request = ""

    for line in lines:

        # Ignore HTTP headers
        if (
            line.startswith("POST ")
            or line.startswith("GET ")
            or line.startswith("Request URL:")
        ):
            continue

        # New request starts whenever we encounter a GA endpoint
        if "google-analytics.com/g/collect" in line or "analytics.google.com/g/collect?" in line:

            # Save previous request
            if current_request:
                requests.append(current_request)

            current_request = line

        else:
            # Continuation of previous request
            if current_request:

                if current_request.endswith("&") or line.startswith("&"):
                    current_request += line
                else:
                    current_request += "&" + line

    if current_request:
        requests.append(current_request)

    # ---------------------------------------
    # Parse each request
    # ---------------------------------------

    for request in requests:

        decoded = urllib.parse.unquote(request)

        # Remove everything before '?'
        if "?" in decoded:
            decoded = decoded.split("?", 1)[1]

        parameters = {}

        for item in decoded.split("&"):

            if "=" not in item:
                continue

            key, value = item.split("=", 1)

            key = key.replace("ep.", "")
            key = key.replace("up.", "")

            parameters[key] = value

        event_name = parameters.get("en", "Unknown")

        parameters["Event Name"] = event_name

        parsed_requests.append(parameters)

    return parsed_requests

def build_tables(parsed_requests):
    """
    Create one dataframe per event type.
    """

    grouped = {}

    for request in parsed_requests:

        event_name = request["Event Name"]

        grouped.setdefault(event_name, []).append(request)

    tables = {}

    for event_name, rows in grouped.items():

        df = pd.DataFrame(rows)

        ordered_columns = []

        # Event Name first
        ordered_columns.append("Event Name")

        # click_id second (if exists)
        if "click_id_hit" in df.columns:
            ordered_columns.append("click_id_hit")
        elif "click_id" in df.columns:
            ordered_columns.append("click_id")

        # Everything else alphabetically
        remaining = sorted(
            [
                c
                for c in df.columns
                if c not in ordered_columns
            ]
        )

        ordered_columns.extend(remaining)

        tables[event_name] = df[ordered_columns]

    return tables

def create_excel(tables):
    """
    Export all event tables into a single worksheet.

    Each event type is written as a separate section:
        PAGE_VIEW
        -----------------
        <table>

        CLICK
        -----------------
        <table>

        FORM_SUBMIT
        -----------------
        <table>
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        workbook = writer.book
        worksheet = workbook.create_sheet("GA4 Requests")

        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        current_row = 1

        for event_name in sorted(tables.keys()):

            df = tables[event_name]

            # Section title
            title_cell = worksheet.cell(row=current_row, column=1)
            title_cell.value = event_name.upper()
            title_cell.font = Font(bold=True, size=14)

            current_row += 1

            # Column headers
            for col_num, column in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = column
                cell.font = Font(bold=True)

            current_row += 1

            # Data
            for row in df.itertuples(index=False):
                for col_num, value in enumerate(row, start=1):
                    worksheet.cell(
                        row=current_row,
                        column=col_num,
                        value=value,
                    )
                current_row += 1

            # Blank rows between sections
            current_row += 3

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(length + 3, 60)

        # Freeze the first row (optional)
        worksheet.freeze_panes = "A2"

    output.seek(0)

    return output


# -----------------------------------------------------------------------------
# UI
# -----------------------------------------------------------------------------

requests = st.text_area(
    "Paste GA4 Requests",
    height=300,
)

if st.button("Process Requests"):

    parsed_requests = parse_requests(requests)

    if len(parsed_requests) == 0:

        st.warning("No valid requests detected.")

    else:

        st.success(f"{len(parsed_requests)} request(s) processed.")

        tables = build_tables(parsed_requests)

        st.divider()

        for event_name in sorted(tables.keys()):

            st.subheader(f"{event_name}")

            st.dataframe(
                tables[event_name],
                use_container_width=True,
                hide_index=True,
            )

        excel = create_excel(tables)

        st.divider()

        st.download_button(
            label="📥 Download Excel",
            data=excel,
            file_name="GA4_Request_Breakdown.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )