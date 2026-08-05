"""
report.py

Generates Excel and HTML QA reports.
"""

from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import REPORT_DIR
from utils import filename_timestamp



############################################################
# Convert results to dataframe
############################################################

def results_to_dataframe(results):

    rows = []

    for r in results:

        rows.append({

            "Test URL":
                r.get("url_tested", ""),

            "Current URL":
                r.get("current_url", ""),

            "Step":
                r.get("step", ""),

            "Action":
                r.get("action", ""),

            "Expected":
                r.get("expected", ""),

            "Actual":
                r.get("actual", ""),

            "OnetrustActiveGroups":
                r.get("active_groups", ""),

            "GA4 Event(s)":
                r.get("ga_event", ""),

            "GA4 Request Count":
                r.get("ga_request_count", 0),

            "Status":
                r.get("status", ""),

            "Notes":
                r.get("notes", ""),

            "Timestamp":
                r.get("timestamp", "")

        })

    return pd.DataFrame(rows)


def ga4_dataframe(results):

    rows = []

    for result in results:

        requests = result.get("ga_requests", [])

        for req in requests:

            rows.append({

                "Test URL":
                    result.get("url_tested", ""),

                "Current URL":
                    result.get("current_url", ""),

                "QA Step":
                    result.get("step", ""),

                "Timestamp":
                    req.get("timestamp", ""),

                "Event":
                    req.get("event", ""),

                "Measurement ID":
                    req.get("measurement_id", ""),

                "Client ID":
                    req.get("client_id", ""),

                "Session ID":
                    req.get("session_id", ""),

                "Page URL":
                    req.get("page_url", ""),

                "Page Title":
                    req.get("page_title", ""),

                "Referrer":
                    req.get("referrer", ""),

                "Consent GCS":
                    req.get("consent_gcs", ""),

                "Consent GCD":
                    req.get("consent_gcd", ""),

                "Full Request":
                    req.get("url", "")

            })

    return pd.DataFrame(rows)

############################################################
# Summary
############################################################


def create_summary(results):

    df = results_to_dataframe(results)


    summary = {

    "URLs Tested":

        df["Test URL"].nunique(),

    "Total Steps":

        len(df),

    "Passed":

        len(
            df[
                df["Status"]=="PASS"
            ]
        ),

    "Failed":

        len(
            df[
                df["Status"]=="FAIL"
            ]
        ),

    "Pass Rate":

        round(
            len(
                df[
                    df["Status"]=="PASS"
                ]
            )
            /
            len(df)
            *
            100,
            2
        )

        if len(df)

        else 0

    }


    return pd.DataFrame(
        [
            summary
        ]
    )



############################################################
# Excel report
############################################################


def generate_excel(results):


    timestamp = filename_timestamp()


    filepath = (
        REPORT_DIR
        /
        f"OneTrust_QA_Report_{timestamp}.xlsx"
    )


    summary = create_summary(results)

    df = results_to_dataframe(results)

    ga_df = ga4_dataframe(results)



    with pd.ExcelWriter(
        filepath,
        engine="openpyxl"
    ) as writer:


        summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )


        df.to_excel(
            writer,
            sheet_name="QA Steps",
            index=False
        )

        ga_df.to_excel(

            writer,

            sheet_name="GA4 Requests",

            index=False

        )



        workbook = writer.book


        for sheet in workbook:


            for column in sheet.columns:


                max_length = 0


                column_letter = (
                    get_column_letter(
                        column[0].column
                    )
                )


                for cell in column:


                    if cell.value:

                        max_length = max(
                            max_length,
                            len(
                                str(cell.value)
                            )
                        )


                sheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 5,
                    50
                )


            for cell in sheet[1]:

                cell.font = Font(
                    bold=True
                )


                cell.alignment = Alignment(
                    horizontal="center"
                )


    return filepath



############################################################
# HTML report
############################################################

def generate_html(results):

    timestamp = filename_timestamp()

    filepath = (
        REPORT_DIR
        /
        f"OneTrust_QA_Report_{timestamp}.html"
    )

    df = results_to_dataframe(results)


    html = f"""
<html>

<head>

<title>
OneTrust Consent QA Report
</title>


<style>

body {{

font-family: Arial;

}}


table {{

border-collapse: collapse;

width:100%;

}}


th {{

background:#333;

color:white;

padding:8px;

}}


td {{

border:1px solid #ddd;

padding:8px;

}}


</style>


</head>


<body>


<h1>
OneTrust Consent Mode QA Report
</h1>


<p>
Generated:
{datetime.now()}
</p>


"""



    html += df.to_html(
        index=False,
        classes="table"
    )

    ga_df = ga4_dataframe(results)

    html += "<br><br>"

    html += "<h2>GA4 Requests</h2>"

    html += ga_df.to_html(
        index=False
    )



    html += """

</body>

</html>

"""


    with open(
        filepath,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(html)


    return filepath


############################################################
# Generate all reports
############################################################


def generate_reports(results):


    outputs = {}


    outputs["excel"] = generate_excel(
        results
    )


    outputs["html"] = generate_html(
        results
    )


    return outputs